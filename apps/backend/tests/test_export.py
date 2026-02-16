"""Unit tests for report export functionality (DOCX, XLSX, PDF)."""

import json

import pytest

from grc_backend.services.report_generator import (
    ExportFormat,
    GeneratedReport,
    ReportGeneratorService,
    ReportType,
)


def _make_report(report_type: ReportType) -> GeneratedReport:
    """Helper: generate a mock report of the given type."""
    from datetime import datetime

    service = ReportGeneratorService(ai_provider=None)
    content = service._generate_mock_content(report_type)
    return GeneratedReport(
        report_type=report_type,
        title=content.get("title", "Test Report"),
        generated_at=datetime.utcnow(),
        content=content,
    )


class TestWordExport:
    """Word (DOCX) export tests."""

    @pytest.fixture
    def service(self):
        return ReportGeneratorService(ai_provider=None)

    @pytest.mark.asyncio
    async def test_summary_to_docx(self, service):
        """SUMMARY report exports to valid DOCX bytes."""
        report = _make_report(ReportType.SUMMARY)
        data = await service.export_report(report, ExportFormat.WORD)
        assert isinstance(data, bytes)
        # DOCX files start with PK (zip signature)
        assert data[:2] == b"PK"

    @pytest.mark.asyncio
    async def test_process_doc_to_docx(self, service):
        """PROCESS_DOC report exports to valid DOCX."""
        report = _make_report(ReportType.PROCESS_DOC)
        data = await service.export_report(report, ExportFormat.WORD)
        assert data[:2] == b"PK"

    @pytest.mark.asyncio
    async def test_rcm_to_docx(self, service):
        """RCM report with table exports to valid DOCX."""
        report = _make_report(ReportType.RCM)
        data = await service.export_report(report, ExportFormat.WORD)
        assert data[:2] == b"PK"

    @pytest.mark.asyncio
    async def test_audit_workpaper_to_docx(self, service):
        """AUDIT_WORKPAPER report exports to valid DOCX."""
        report = _make_report(ReportType.AUDIT_WORKPAPER)
        data = await service.export_report(report, ExportFormat.WORD)
        assert data[:2] == b"PK"


class TestExcelExport:
    """Excel (XLSX) export tests."""

    @pytest.fixture
    def service(self):
        return ReportGeneratorService(ai_provider=None)

    @pytest.mark.asyncio
    async def test_rcm_to_xlsx(self, service):
        """RCM report exports to valid XLSX with table data."""
        report = _make_report(ReportType.RCM)
        data = await service.export_report(report, ExportFormat.EXCEL)
        assert isinstance(data, bytes)
        assert data[:2] == b"PK"  # XLSX is also a zip

    @pytest.mark.asyncio
    async def test_process_doc_to_xlsx(self, service):
        """PROCESS_DOC exports to XLSX with process steps."""
        report = _make_report(ReportType.PROCESS_DOC)
        data = await service.export_report(report, ExportFormat.EXCEL)
        assert data[:2] == b"PK"

    @pytest.mark.asyncio
    async def test_summary_to_xlsx_generic(self, service):
        """SUMMARY (no table structure) exports as generic key-value XLSX."""
        report = _make_report(ReportType.SUMMARY)
        data = await service.export_report(report, ExportFormat.EXCEL)
        assert data[:2] == b"PK"

    @pytest.mark.asyncio
    async def test_xlsx_can_be_opened(self, service):
        """Exported XLSX can be opened by openpyxl."""
        from io import BytesIO

        from openpyxl import load_workbook

        report = _make_report(ReportType.RCM)
        data = await service.export_report(report, ExportFormat.EXCEL)
        wb = load_workbook(BytesIO(data))
        ws = wb.active
        # Header row should exist
        assert ws.cell(row=1, column=1).value is not None
        # Data row should exist
        assert ws.cell(row=2, column=1).value is not None


class TestPDFExport:
    """PDF export tests."""

    @pytest.fixture
    def service(self):
        return ReportGeneratorService(ai_provider=None)

    @pytest.mark.asyncio
    async def test_summary_to_pdf(self, service):
        """SUMMARY exports to valid PDF bytes."""
        report = _make_report(ReportType.SUMMARY)
        data = await service.export_report(report, ExportFormat.PDF)
        assert isinstance(data, bytes)
        # PDF starts with %PDF
        assert data[:5] == b"%PDF-"

    @pytest.mark.asyncio
    async def test_rcm_to_pdf_with_table(self, service):
        """RCM report exports to PDF with table."""
        report = _make_report(ReportType.RCM)
        data = await service.export_report(report, ExportFormat.PDF)
        assert data[:5] == b"%PDF-"

    @pytest.mark.asyncio
    async def test_audit_workpaper_to_pdf(self, service):
        """AUDIT_WORKPAPER exports to PDF."""
        report = _make_report(ReportType.AUDIT_WORKPAPER)
        data = await service.export_report(report, ExportFormat.PDF)
        assert data[:5] == b"%PDF-"

    @pytest.mark.asyncio
    async def test_process_doc_to_pdf(self, service):
        """PROCESS_DOC exports to PDF."""
        report = _make_report(ReportType.PROCESS_DOC)
        data = await service.export_report(report, ExportFormat.PDF)
        assert data[:5] == b"%PDF-"


class TestExportEdgeCases:
    """Edge case tests for export."""

    @pytest.fixture
    def service(self):
        return ReportGeneratorService(ai_provider=None)

    @pytest.mark.asyncio
    async def test_unsupported_format_raises(self, service):
        """Unsupported format raises ValueError."""
        report = _make_report(ReportType.SUMMARY)
        with pytest.raises(ValueError, match="Unsupported export format"):
            await service.export_report(report, "invalid_format")

    @pytest.mark.asyncio
    async def test_empty_content_docx(self, service):
        """Export with minimal/empty content does not crash."""
        from datetime import datetime

        report = GeneratedReport(
            report_type=ReportType.SUMMARY,
            title="Empty",
            generated_at=datetime.utcnow(),
            content={"title": "Empty Report"},
        )
        data = await service.export_report(report, ExportFormat.WORD)
        assert data[:2] == b"PK"

    @pytest.mark.asyncio
    async def test_empty_content_xlsx(self, service):
        """Export XLSX with minimal content produces valid file."""
        from datetime import datetime

        report = GeneratedReport(
            report_type=ReportType.RCM,
            title="Empty RCM",
            generated_at=datetime.utcnow(),
            content={"title": "Empty RCM", "items": []},
        )
        data = await service.export_report(report, ExportFormat.EXCEL)
        assert data[:2] == b"PK"

    @pytest.mark.asyncio
    async def test_empty_content_pdf(self, service):
        """Export PDF with minimal content produces valid file."""
        from datetime import datetime

        report = GeneratedReport(
            report_type=ReportType.SUMMARY,
            title="Empty",
            generated_at=datetime.utcnow(),
            content={"title": "Empty Report"},
        )
        data = await service.export_report(report, ExportFormat.PDF)
        assert data[:5] == b"%PDF-"

    @pytest.mark.asyncio
    async def test_json_export_roundtrip(self, service):
        """JSON export can be parsed back."""
        report = _make_report(ReportType.SUMMARY)
        data = await service.export_report(report, ExportFormat.JSON)
        parsed = json.loads(data.decode("utf-8"))
        assert parsed["title"] == report.content["title"]
